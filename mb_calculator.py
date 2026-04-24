from __future__ import annotations

import argparse
import copy
import math
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel, to_excel


BASE_DIR = Path(__file__).resolve().parent
SHEET_NAME = "MBP_CPP_CHP_RFP_SMB_MSB_RMM_MMP"
INPUT_LAST_COL = 56  # BD
CALC_FIRST_COL = 58  # BF
CALC_LAST_COL = 119  # DO
DATA_START_ROW = 3
CURRENT_YEAR = 2026
ACCRUED_INTEREST_DATE_SERIAL = 46112  # 2026-03-31 in Excel's date system
ACCRUED_INTEREST_RATE = 0.0825
FORMAT_TEMPLATE_PATH = BASE_DIR / "format_template.xlsx"


ROW1_HEADERS = {
    "BF": "Protiviti Calc",
    "BJ": "To Update >>>",
    "BL": ACCRUED_INTEREST_DATE_SERIAL,
    "BM": ACCRUED_INTEREST_RATE,
    "BN": "INFORCE/PAID UP DEATH",
    "BP": "CHP",
    "BX": " ",
    "BY": "MBP",
    "CE": "CPP and CHP",
    "CF": "CPP",
    "CG": "CHP",
    "CH": "CPP \n(A)",
    "CM": "CHP",
    "CR": "RFP",
    "CX": "RMM, MMP",
    "CY": "RMM, MMP",
    "CZ": "RMM and MMP has Loan",
    "DC": "SMB",
    "DF": "XXXX",
}


ROW2_HEADERS = {
    "BF": "Inforce/Paid up",
    "BG": "PPT * FREQ",
    "BH": "Prem Paid * FREQ",
    "BI": "PU Factor",
    "BJ": "Annualised Premium\nCHP",
    "BK": "Updated SA\n(PU/IF)",
    "BL": "For Current Month Accrued Loan Interest",
    "BM": "Updated LN_LA",
    "BN": "Updated SA \n(Not used)",
    "BO": "Death Case\n(Inforce, paid up or Lapsed)",
    "BP": "Updated SA PU/IF\n(BSA + HSA)\nfor CHP",
    "BQ": "ORIGINAL_SA +HSA for CHP",
    "BR": "Premium paid (Years)",
    "BS": "Policy Lapse before death",
    "BT": "Current Year",
    "BU": "Policy Start Year \n(Special Bonus)",
    "BV": "MBP\nSB Payable\n(Times)",
    "BW": "MBP\nSB Paid \n(Times)",
    "BX": "MBP\nSB Paid",
    "BY": "Survival Benefit on Maturity\n(A)",
    "BZ": "MBP\nRV Bonus\n(B)",
    "CA": "Special Bonus % (MBP)",
    "CB": "Special Bonus Amt \n(MBP)",
    "CC": "Total \nMBP Maturity",
    "CD": "CPP, CHP SB PBFable Eligibility",
    "CE": "After how many installments has been paid, policy has become paid-up (CPP & CHP)",
    "CF": "SB Amount paid till 3 installments\nCPP",
    "CG": "SB Amount paid till 3 installments\nCHP",
    "CH": "CPP RV Bonus \n(A)",
    "CI": "CPP Special Bonus %",
    "CJ": "CPP Special Bonus Amount (B)",
    "CK": "CPP SB on Maturity\n C",
    "CL": "TOTAL \nCPP MB\n(A+B+C)",
    "CM": "High Sum Assured\nCHP",
    "CN": "CHP SB on Maturity (Base Benefit)",
    "CO": "CHP RV Bonus (LE RY)",
    "CP": "CHP Bonus (LE SI)*",
    "CQ": "Total maturity for CHP",
    "CR": "SA Due or Rev. Bonus Due (RFP)",
    "CS": "RFP RV Bonus",
    "CT": "Special Bonus% ",
    "CU": "Special Bonus Amt",
    "CV": "Basic Benefit",
    "CW": "Total MB Payable for RFP",
    "CX": "GLA\n(RMM, MMP)",
    "CY": "GLA\n(RMM, MMP)",
    "CZ": "GMA\n(RMM. MMP)",
    "DA": "Basic Benefit",
    "DB": "Total RMM, MMP",
    "DC": "Survival Benefit SMB, MSB",
    "DD": "GMA \nSMB",
    "DE": "TOTAL MB for SMB",
    "DF": "TOTAL Maturity Benefit",
    "DG": "LN_LA \nUpdated",
    "DH": "GLA",
    "DI": "GMA",
    "DJ": "Revisionary Bonus",
    "DK": "Special Bonus",
    "DL": "Terminal Bonus",
    "DM": "Maturity Benefit",
    "DN": "Condition of 100.1% , additional bonus",
    "DO": "Total Maturity Benefit ",
}


MBP_RV_RATES = {
    2001: 0.0, 2002: 0.03, 2003: 0.035, 2004: 0.03, 2005: 0.0275,
    2006: 0.03, 2007: 0.03, 2008: 0.03, 2009: 0.026, 2010: 0.026,
    2011: 0.026, 2012: 0.023, 2013: 0.023, 2014: 0.023, 2015: 0.021,
    2016: 0.0225, 2017: 0.025, 2018: 0.0255, 2019: 0.0295, 2020: 0.0295,
    2021: 0.0325, 2022: 0.0333, 2023: 0.035, 2024: 0.035, 2025: 0.035,
    2026: 0.035,
}

CPP_RV_RATES = {
    2001: 0.0, 2002: 0.03, 2003: 0.035, 2004: 0.03, 2005: 0.0275,
    2006: 0.03, 2007: 0.03, 2008: 0.03, 2009: 0.026, 2010: 0.026,
    2011: 0.026, 2012: 0.023, 2013: 0.023, 2014: 0.023, 2015: 0.021,
    2016: 0.025, 2017: 0.025, 2018: 0.0295, 2019: 0.032, 2020: 0.032,
    2021: 0.034, 2022: 0.038, 2023: 0.0423, 2024: 0.0423, 2025: 0.0423,
    2026: 0.0423,
}

CHP_RV_RATES = {
    2001: 0.0, 2002: 0.0, 2003: 0.0, 2004: 0.0, 2005: 0.0, 2006: 0.0,
    2007: 0.0, 2008: 0.0, 2009: 0.0, 2010: 0.0, 2011: 0.0, 2012: 0.0,
    2013: 0.0, 2014: 0.023, 2015: 0.021, 2016: 0.025, 2017: 0.025,
    2018: 0.025, 2019: 0.03, 2020: 0.03, 2021: 0.03, 2022: 0.03,
    2023: 0.03, 2024: 0.03, 2025: 0.03, 2026: 0.03,
}

RFP_RV_RATES = {
    2001: 0.0, 2002: 0.0, 2003: 0.0, 2004: 0.0, 2005: 0.0, 2006: 0.0,
    2007: 0.0, 2008: 0.0, 2009: 0.0, 2010: 0.026, 2011: 0.026,
    2012: 0.023, 2013: 0.023, 2014: 0.023, 2015: 0.021, 2016: 0.021,
    2017: 0.021, 2018: 0.023, 2019: 0.03, 2020: 0.03, 2021: 0.032,
    2022: 0.032, 2023: 0.032, 2024: 0.032, 2025: 0.032, 2026: 0.032,
}

SPECIAL_BONUS_RATES = {
    # Entry year: (MBP, CPP, RFP)
    2001: (0.0, 0.0, 0.0),
    2002: (0.0, 0.0, 0.0),
    2003: (0.0052, 0.0, 0.0),
    2004: (0.0, 0.0, 0.0),
    2005: (0.0, 0.0123, 0.0),
    2006: (0.0054, 0.0123, 0.0),
    2007: (0.0, 0.0123, 0.0),
    2008: (0.0, 0.0123, 0.0),
    2009: (0.0055, 0.012, 0.0083),
    2010: (0.0, 0.0117, 0.008),
    2011: (0.0, 0.0114, 0.0077),
    2012: (0.0049, 0.0108, 0.007),
    2013: (0.0, 0.0101, 0.0064),
    2014: (0.0, 0.0, 0.0),
    2015: (0.0, 0.0, 0.0),
}

MBP_SB_PAID_TIMES = {
    1: [(0, 0), (4, 1), (7, 2), (10, 3), (13, 4), (16, 5), (19, 6)],
    2: [(0, 0), (8, 1), (14, 2), (20, 3), (26, 4), (32, 5), (38, 6)],
    4: [(0, 0), (16, 1), (28, 2), (40, 3), (52, 4), (64, 5), (76, 6)],
    12: [(0, 0), (48, 1), (84, 2), (120, 3), (156, 4), (192, 5), (228, 6)],
}

RMM_GLA_RATES = {
    1: 0.01, 2: 0.03, 3: 0.06, 4: 0.10, 5: 0.15, 6: 0.21, 7: 0.28,
    8: 0.36, 9: 0.45, 10: 0.55, 11: 0.66, 12: 0.78, 13: 0.91, 14: 1.05,
    15: 1.20, 16: 1.36, 17: 1.53, 18: 1.71, 19: 1.90, 20: 2.10,
}


@dataclass
class CalculationRow:
    values: dict[str, Any]

    def get(self, col: str) -> Any:
        return self.values.get(col)

    def set(self, col: str, value: Any) -> Any:
        self.values[col] = value
        return value


def col_num(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + ord(ch.upper()) - 64
    return n


def is_blank(value: Any) -> bool:
    return value is None or value == ""


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def eq_text(left: Any, right: str) -> bool:
    return clean_text(left).casefold() == right.casefold()


def is_one_of(value: Any, *options: str) -> bool:
    return any(eq_text(value, option) for option in options)


def num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime):
        return float(to_excel(value))
    if isinstance(value, date):
        return float(to_excel(datetime(value.year, value.month, value.day)))
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return 0.0
        return float(stripped)
    return float(value)


def safe_num(value: Any) -> float:
    try:
        return num(value)
    except Exception:
        return 0.0


def excel_sum(*values: Any) -> float:
    total = 0.0
    for value in values:
        if isinstance(value, (list, tuple)):
            total += excel_sum(*value)
        else:
            total += safe_num(value)
    return total


def excel_round(value: Any, digits: int = 0) -> float:
    q = Decimal("1").scaleb(-digits)
    rounded = Decimal(str(num(value))).quantize(q, rounding=ROUND_HALF_UP)
    result = float(rounded)
    return int(result) if digits == 0 else result


def rounddown(value: Any, digits: int = 0) -> float:
    factor = 10 ** digits
    value_num = num(value) * factor
    result = math.floor(value_num) / factor if value_num >= 0 else math.ceil(value_num) / factor
    return int(result) if digits == 0 else result


def excel_date(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        return from_excel(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(stripped, fmt)
            except ValueError:
                pass
        return from_excel(float(stripped))
    return None


def date_serial(value: Any) -> float:
    dt = excel_date(value)
    if dt is None:
        return 0.0
    return float(to_excel(dt))


def date_add_days(value: Any, days: int) -> float:
    return date_serial(value) + days


def datedif_years(start: Any, end: Any) -> int:
    start_dt = excel_date(start)
    end_dt = excel_date(end)
    if start_dt is None or end_dt is None:
        return 0
    years = end_dt.year - start_dt.year
    if (end_dt.month, end_dt.day) < (start_dt.month, start_dt.day):
        years -= 1
    return years


def edate(value: Any, months: int) -> datetime | None:
    dt = excel_date(value)
    if dt is None:
        return None
    month_index = dt.month - 1 + months
    year = dt.year + month_index // 12
    month = month_index % 12 + 1
    last_day = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28,
                31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1]
    return datetime(year, month, min(dt.day, last_day))


def fiscal_year_from_rcd(value: Any) -> int:
    dt = excel_date(value)
    if dt is None:
        return 0
    return dt.year if dt.month <= 3 else dt.year + 1


def paid_to_fiscal_year(paid_to_date: Any, death_status: Any | None = None) -> int | str:
    dt = excel_date(paid_to_date)
    if dt is None:
        return ""
    if death_status is not None and eq_text(death_status, "Inforce death"):
        return CURRENT_YEAR
    return dt.year - 1 if dt.month <= 3 else dt.year


def special_bonus_rate(entry_year: Any, product: str) -> float:
    rates = SPECIAL_BONUS_RATES.get(int(num(entry_year)), (0.0, 0.0, 0.0))
    index = {"MBP": 0, "CPP": 1, "RFP": 2}[product]
    return rates[index] or 0.0


def mbp_sb_paid_times(frequency: Any, premium_paid_x_frequency: Any) -> Any:
    table = MBP_SB_PAID_TIMES.get(int(num(frequency)))
    if not table:
        return "NA"
    paid = num(premium_paid_x_frequency)
    result = 0
    for threshold, benefit in table:
        if paid >= threshold:
            result = benefit
        else:
            break
    return result


def rmm_gla_rate(years: Any) -> float:
    return RMM_GLA_RATES.get(int(num(years)), 0.0)


def simple_rv_bonus(row: CalculationRow, rates: dict[int, float], death_status: Any | None = None) -> float:
    rcd_fy = fiscal_year_from_rcd(row.get("AF"))
    ptd_fy = paid_to_fiscal_year(row.get("AG"), death_status)
    if ptd_fy == "":
        return 0.0
    original_sa = num(row.get("AT"))
    total = 0.0
    for year in range(2001, CURRENT_YEAR + 1):
        if rcd_fy <= year and int(ptd_fy) >= year:
            total += original_sa * rates.get(year, 0.0)
    return total


def rfp_rv_bonus(row: CalculationRow) -> float:
    rcd_fy = fiscal_year_from_rcd(row.get("AF"))
    ptd_fy = paid_to_fiscal_year(row.get("AG"))
    if ptd_fy == "":
        return 0.0
    ptd_final = CURRENT_YEAR if int(ptd_fy) + 5 == CURRENT_YEAR else int(ptd_fy)
    running_base = num(row.get("AT"))
    total_bonus = 0.0
    for year in range(2001, CURRENT_YEAR + 1):
        if rcd_fy <= year and ptd_final >= year:
            bonus = running_base * RFP_RV_RATES.get(year, 0.0)
            total_bonus += bonus
            running_base += bonus
    return total_bonus


def calculate_row(input_values: dict[str, Any]) -> CalculationRow:
    row = CalculationRow(dict(input_values))

    bf = row.set("BF", row.get("B"))
    bg = row.set("BG", num(row.get("AD")) * num(row.get("AE")))
    bh = row.set("BH", excel_round((date_serial(row.get("AG")) - date_serial(row.get("AF"))) / 365 * num(row.get("AE")), 0))
    bi = row.set("BI", num(bh) / num(bg) if num(bg) else 0)
    bj = row.set("BJ", excel_round(
        num(row.get("BB")) if eq_text(row.get("Z"), "CHP") and num(row.get("AE")) == 1 else
        num(row.get("BB")) / 0.51 if eq_text(row.get("Z"), "CHP") and num(row.get("AE")) == 2 else
        num(row.get("BB")) / 0.26 if eq_text(row.get("Z"), "CHP") and num(row.get("AE")) == 4 else
        num(row.get("BB")) / 0.0834 if eq_text(row.get("Z"), "CHP") and num(row.get("AE")) == 12 else
        0,
        0,
    ))

    bo = "NA"
    if eq_text(row.get("B"), "DH"):
        grace_days = 30 if num(row.get("AE")) != 12 else 15
        bo = "Inforce Death" if date_add_days(row.get("AG"), grace_days) >= date_serial(row.get("AK")) else "Paidup Death"
    row.set("BO", bo)

    bk = 0.0
    if is_one_of(bf, "IF", "PM") or eq_text(bo, "Inforce Death"):
        bk = num(row.get("AT"))
    elif eq_text(bf, "PU") or eq_text(bo, "Paidup Death"):
        bk = num(row.get("AT")) * num(bh) / num(bg) if num(bg) else 0.0
    row.set("BK", bk)

    ah_serial = date_serial(row.get("AH"))
    bl = 0.0 if ah_serial <= ACCRUED_INTEREST_DATE_SERIAL else (
        (num(row.get("M")) + num(row.get("N"))) *
        ((1 + ACCRUED_INTEREST_RATE) ** ((ah_serial - ACCRUED_INTEREST_DATE_SERIAL) / 365) - 1)
    )
    row.set("BL", bl)
    row.set("BM", num(row.get("N")) + num(bl))

    bn = 0.0
    if is_one_of(row.get("Z"), "CPP", "CHP"):
        if is_one_of(bf, "IF", "PM") or eq_text(bo, "Inforce Death"):
            bn = num(row.get("AT"))
        elif eq_text(bf, "PU") or eq_text(bo, "Paidup Death"):
            bn = num(row.get("AT")) * num(bh) / num(bg) if num(bg) else 0.0
    row.set("BN", bn)

    br = row.set("BR", rounddown(datedif_years(row.get("AF"), row.get("AG")), 0))
    bs = "NA"
    if eq_text(row.get("B"), "DH"):
        grace_days = 30 if num(row.get("AE")) != 12 else 15
        if num(br) < 3 and date_add_days(row.get("AG"), grace_days) < date_serial(row.get("AK")):
            bs = "Lapsed Death"
        elif date_add_days(row.get("AG"), grace_days) >= date_serial(row.get("AK")):
            bs = "Inforce Death"
        else:
            bs = "Paidup Death"
    row.set("BS", bs)

    cm = (0 if not eq_text(row.get("Z"), "CHP") else
          0 if num(row.get("AT")) < 250000 else
          num(row.get("AT")) * 0.002 if num(row.get("AT")) < 500000 else
          num(row.get("AT")) * 0.003) / 4
    row.set("CM", cm)

    bq = row.set("BQ", num(row.get("AT")) + (num(cm) * 4) if eq_text(row.get("Z"), "CHP") else 0)
    bp = 0.0
    if eq_text(row.get("Z"), "CHP"):
        if is_one_of(bf, "IF", "PM") or eq_text(bo, "Inforce Death"):
            bp = num(bq)
        elif eq_text(bf, "PU") or eq_text(bo, "Paidup Death"):
            bp = num(bq) * num(bh) / num(bg) if num(bg) else 0.0
    row.set("BP", bp)

    row.set("BT", datedif_years(row.get("AF"), row.get("AH")))
    edate_value = edate(row.get("AF"), -3)
    bu = row.set("BU", edate_value.year if edate_value is not None else 0)

    bv = row.set("BV", (num(row.get("AC")) - 1) / 3 if eq_text(row.get("Z"), "MBP") else "0")
    bw = row.set("BW", mbp_sb_paid_times(row.get("AE"), row.get("BH")))
    bx = "0"
    if eq_text(row.get("Z"), "MBP"):
        if num(bv) != num(bw):
            bx = num(row.get("AT")) / num(bv) * num(bw) if num(bv) else 0.0
        else:
            bx = num(row.get("AT")) / num(bv) * (num(bw) - 1) if num(bv) else 0.0
    row.set("BX", bx)
    by = row.set("BY", num(bk) - num(bx) if eq_text(row.get("Z"), "MBP") else 0)
    bz = row.set("BZ", simple_rv_bonus(row, MBP_RV_RATES) if eq_text(row.get("Z"), "MBP") else 0)
    ca = row.set("CA", special_bonus_rate(bu, "MBP") if eq_text(row.get("Z"), "MBP") and eq_text(bf, "IF") else "0")
    try:
        cb = num(bk) * num(ca) * num(row.get("AC"))
    except Exception:
        cb = "0"
    row.set("CB", cb)
    row.set("CC", num(by) + num(bz) + num(cb) if eq_text(row.get("Z"), "MBP") else "0")

    row.set("CD", "SB DUE" if is_one_of(row.get("Z"), "CPP", "CHP") and num(row.get("AC")) - num(row.get("BT")) <= 4
            else "SB NOT DUE" if is_one_of(row.get("Z"), "CPP", "CHP") else 0)
    ce = row.set("CE", rounddown(min(max(num(bi) * num(row.get("AC")) - (num(row.get("AC")) - 4), 0), 3), 0))

    cf = 0.0
    if eq_text(row.get("Z"), "CPP"):
        if eq_text(bo, "Inforce Death"):
            cf = num(row.get("AT")) / 4 * 3
        elif num(ce) == 0:
            cf = (num(row.get("AT")) * num(bi) / 4) * 3
        elif num(ce) >= 3:
            cf = num(row.get("AT")) / 4 * 3
        else:
            cf = (num(row.get("AT")) / 4 * num(ce) +
                  ((num(row.get("AT")) * num(bi) - num(row.get("AT")) / 4 * num(ce)) / (4 - num(ce))) *
                  min(3 - num(ce), 3))
    row.set("CF", cf)

    cg = 0.0
    if eq_text(row.get("Z"), "CHP"):
        if eq_text(bo, "Inforce Death"):
            cg = num(bq) / 4 * 3
        elif num(ce) == 0:
            cg = (num(bq) * num(bi) / 4) * 3
        elif num(ce) >= 3:
            cg = num(bq) / 4 * 3
        else:
            cg = num(bq) / 4 * num(ce) + ((num(bq) * num(bi) * 0.25) * min(3 - num(ce), 3))
    row.set("CG", cg)

    ch = row.set("CH", simple_rv_bonus(row, CPP_RV_RATES, bo) if eq_text(row.get("Z"), "CPP") else "0")
    ci = row.set("CI", special_bonus_rate(bu, "CPP") if eq_text(row.get("Z"), "CPP") and eq_text(bf, "IF") else "0")
    cj = row.set("CJ", num(bk) * num(ci) * num(row.get("AC")) if eq_text(row.get("Z"), "CPP") else "0")
    ck = row.set("CK", num(bk) - num(cf) if eq_text(row.get("Z"), "CPP") else "0")
    row.set("CL", num(ch) + num(cj) + num(ck) if eq_text(row.get("Z"), "CPP") else "0")

    cn = 0.0
    if eq_text(row.get("Z"), "CHP"):
        if eq_text(bf, "IF") or eq_text(bo, "Inforce Death"):
            cn = num(bq) * 0.25
        elif eq_text(bf, "PU") or eq_text(bo, "Paidup Death"):
            cn = num(bq) * num(bi) * 0.25
    row.set("CN", cn)
    co = row.set("CO", simple_rv_bonus(row, CHP_RV_RATES, bo) if eq_text(row.get("Z"), "CHP") else "0")
    cp = "0"
    if eq_text(row.get("Z"), "CHP"):
        cp = 0
        if eq_text(bf, "IF") or eq_text(bo, "Inforce Death"):
            threshold = num(bj) * num(row.get("AD")) * 1.001
            total_existing = num(cn) + num(co) + num(cg)
            cp = threshold - total_existing if threshold > total_existing else 0
    row.set("CP", cp)
    row.set("CQ", num(cm) + num(cn) + num(cp) + num(co) if eq_text(row.get("Z"), "CHP") else 0)

    cr = "NA"
    if eq_text(row.get("Z"), "RFP"):
        cr = "Rev.Bonus Due" if num(row.get("BT")) == num(row.get("AC")) else (
            "Sum Assured Due" if num(row.get("BT")) == num(row.get("AD")) else "NA"
        )
    row.set("CR", cr)
    cs = row.set("CS", rfp_rv_bonus(row) if eq_text(row.get("Z"), "RFP") and is_one_of(bf, "PM", "IF", "PU") and eq_text(cr, "Rev.Bonus Due") else 0)
    ct = row.set("CT", special_bonus_rate(bu, "RFP") if eq_text(row.get("Z"), "RFP") and num(bg) == num(bh) else "0")
    cu = row.set("CU", num(bk) * num(ct) * num(row.get("AC")))
    cv = row.set("CV", num(bk) if eq_text(row.get("Z"), "RFP") and eq_text(cr, "Sum Assured Due") else 0)
    row.set("CW", num(cs) + num(cu) + num(cv))

    cx = "0"
    if is_one_of(row.get("Z"), "MMP", "RMM") and is_one_of(bf, "IF", "PM", "PU"):
        cx = num(row.get("AT")) * rmm_gla_rate(br)
    row.set("CX", cx)
    row.set("CY", 0)
    cz = "0"
    if is_one_of(row.get("Z"), "RMM", "MMP") and eq_text(bf, "IF"):
        cz = num(row.get("AT")) * 0.01 * num(row.get("AC"))
    row.set("CZ", cz)
    da = row.set("DA", num(bk) if is_one_of(row.get("Z"), "RMM", "MMP") else 0)
    row.set("DB", (num(da) + num(cx)) if is_one_of(row.get("Z"), "RMM", "MMP") and eq_text(bf, "PU")
            else (num(cx) + num(cz) + num(da)) if is_one_of(row.get("Z"), "RMM", "MMP") else "0")

    dc = "0"
    if is_one_of(row.get("Z"), "MSB", "SMB") and is_one_of(bf, "IF", "PU"):
        if num(row.get("BT")) > num(row.get("AC")):
            dc = "NA"
        else:
            dc = ((500 / num(row.get("AC"))) if num(row.get("BT")) % 5 == 0 else 0) / 100 * num(bk)
    row.set("DC", dc)
    dd = row.set("DD", num(bk) * 0.01 * num(row.get("AC")) if eq_text(row.get("Z"), "SMB") and is_one_of(bf, "IF", "PU") else 0)
    row.set("DE", num(bk) - safe_num(dc) + num(dd) if eq_text(row.get("Z"), "SMB") else "0")

    product_total = 0
    if eq_text(row.get("Z"), "MBP"):
        product_total = row.get("CC")
    elif eq_text(row.get("Z"), "CPP"):
        product_total = row.get("CL")
    elif eq_text(row.get("Z"), "CHP"):
        product_total = row.get("CQ")
    elif eq_text(row.get("Z"), "RFP"):
        product_total = row.get("CW")
    elif is_one_of(row.get("Z"), "RMM", "MMP"):
        product_total = row.get("DB")
    elif is_one_of(row.get("Z"), "SMB", "MSB"):
        product_total = row.get("DE")
    row.set("DF", safe_num(product_total) - excel_sum(row.get("M"), row.get("N"), row.get("O")) - safe_num(row.get("Q")))
    row.set("DG", row.get("BM"))
    row.set("DH", row.get("CX"))
    row.set("DI", safe_num(row.get("CZ")) + safe_num(row.get("DD")))
    row.set("DJ", "0" if eq_text(row.get("Z"), "CHP") and eq_text(bf, "DH") else excel_sum(row.get("BZ"), row.get("CH"), row.get("CO"), row.get("CS")))
    row.set("DK", excel_sum(row.get("CB"), row.get("CJ"), row.get("CU")))
    row.set("DL", 0)
    row.set("DM", excel_sum(row.get("BY"), row.get("CK"), row.get("CN"), row.get("DC"), row.get("CV"), row.get("DA")))
    row.set("DN", row.get("CP"))
    row.set("DO", excel_sum([row.get(col) for col in ["DH", "DI", "DJ", "DK", "DL", "DM", "DN"]]) -
            excel_sum(row.get("M"), row.get("DG")) - safe_num(row.get("Q")))
    return row


def build_output(input_path: Path, output_path: Path) -> int:
    input_wb = load_workbook(input_path, data_only=True, read_only=True)
    input_ws = input_wb[SHEET_NAME] if SHEET_NAME in input_wb.sheetnames else input_wb[input_wb.sheetnames[0]]

    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = SHEET_NAME

    rows_written = 0
    data_rows = 0
    for row_num, row_values in enumerate(
        input_ws.iter_rows(min_row=1, max_col=INPUT_LAST_COL, values_only=True),
        start=1,
    ):
        if row_num >= DATA_START_ROW and not any(value is not None for value in row_values):
            continue
        rows_written = max(rows_written, row_num)
        for col, value in enumerate(row_values, start=1):
            output_ws.cell(row_num, col).value = value

    apply_template_formatting(output_wb, output_ws, rows_written)

    for col in range(CALC_FIRST_COL, CALC_LAST_COL + 1):
        letter = get_column_letter(col)
        output_ws.column_dimensions[letter].width = 14
        header_1 = ROW1_HEADERS.get(letter)
        header_2 = ROW2_HEADERS.get(letter)
        cell_1 = output_ws[f"{letter}1"]
        cell_2 = output_ws[f"{letter}2"]
        if header_1 is not None and not isinstance(cell_1, MergedCell):
            cell_1.value = header_1
        if header_2 is not None and not isinstance(cell_2, MergedCell):
            cell_2.value = header_2

    for row_num, row_values in enumerate(
        output_ws.iter_rows(min_row=DATA_START_ROW, max_row=rows_written, max_col=INPUT_LAST_COL, values_only=True),
        start=DATA_START_ROW,
    ):
        if not any(value is not None for value in row_values):
            continue
        data_rows += 1
        input_values = {get_column_letter(col): value for col, value in enumerate(row_values, start=1)}
        calc = calculate_row(input_values)
        for col in range(CALC_FIRST_COL, CALC_LAST_COL + 1):
            letter = get_column_letter(col)
            output_ws.cell(row_num, col).value = calc.get(letter)

    output_ws.freeze_panes = "A3"
    output_wb.save(output_path)
    return data_rows


def copy_cell_format(source_cell: Any, target_cell: Any) -> None:
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format


def apply_template_formatting(
    output_wb: openpyxl.Workbook,
    output_ws: openpyxl.worksheet.worksheet.Worksheet,
    rows_written: int,
) -> None:
    if not FORMAT_TEMPLATE_PATH.exists():
        return

    template_wb = load_workbook(FORMAT_TEMPLATE_PATH)
    template_ws = template_wb[SHEET_NAME]

    for merged_range in template_ws.merged_cells.ranges:
        output_ws.merge_cells(str(merged_range))

    for col in range(1, CALC_LAST_COL + 1):
        letter = get_column_letter(col)
        output_ws.column_dimensions[letter].width = template_ws.column_dimensions[letter].width

    for row_num in (1, 2, 3):
        if template_ws.row_dimensions[row_num].height is not None:
            output_ws.row_dimensions[row_num].height = template_ws.row_dimensions[row_num].height

    template_data_row = 3
    data_row_height = template_ws.row_dimensions[template_data_row].height

    for row_num in range(1, max(rows_written, 2) + 1):
        template_row = row_num if row_num <= 3 else template_data_row
        if row_num >= DATA_START_ROW and data_row_height is not None:
            output_ws.row_dimensions[row_num].height = data_row_height
        for col in range(1, CALC_LAST_COL + 1):
            copy_cell_format(template_ws.cell(template_row, col), output_ws.cell(row_num, col))

    output_ws.freeze_panes = "A3"

    if template_ws.sheet_view.showGridLines is not None:
        output_ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines

    output_ws.sheet_format.defaultColWidth = template_ws.sheet_format.defaultColWidth
    output_ws.sheet_format.defaultRowHeight = template_ws.sheet_format.defaultRowHeight
    output_ws.sheet_format.baseColWidth = template_ws.sheet_format.baseColWidth

    output_ws.page_margins = copy.copy(template_ws.page_margins)
    output_ws.page_setup = copy.copy(template_ws.page_setup)
    output_ws.print_options = copy.copy(template_ws.print_options)


def values_close(left: Any, right: Any, tolerance: float = 1e-6) -> bool:
    if is_blank(left) and is_blank(right):
        return True
    try:
        return abs(num(left) - num(right)) <= tolerance
    except Exception:
        return clean_text(left).casefold() == clean_text(right).casefold()


def validate_output(output_path: Path, reference_path: Path) -> list[str]:
    out_wb = load_workbook(output_path, data_only=True, read_only=True)
    ref_wb = load_workbook(reference_path, data_only=True, read_only=True)
    out_ws = out_wb[SHEET_NAME]
    ref_ws = ref_wb[SHEET_NAME]
    mismatches: list[str] = []
    max_row = out_ws.max_row
    for row_num in range(DATA_START_ROW, max_row + 1):
        if not any(ref_ws.cell(row_num, col).value is not None for col in range(1, INPUT_LAST_COL + 1)):
            continue
        for col in range(CALC_FIRST_COL, CALC_LAST_COL + 1):
            letter = get_column_letter(col)
            expected = ref_ws.cell(row_num, col).value
            actual = out_ws.cell(row_num, col).value
            # The supplied workbook has a manually typed "o" in CR3. The formula used
            # from the next row onward returns NA for non-RFP policies.
            if row_num == 3 and letter == "CR" and expected == "o" and actual == "NA":
                continue
            if not values_close(actual, expected):
                mismatches.append(f"{letter}{row_num}: expected {expected!r}, got {actual!r}")
                if len(mismatches) >= 50:
                    return mismatches
    return mismatches


def main() -> int:
    parser = argparse.ArgumentParser(description="Calculate MB maturity benefit output columns BF:DO from input columns A:BD.")
    parser.add_argument("input", type=Path, help="Input Excel file with one sheet and columns A:BD.")
    parser.add_argument("output", type=Path, help="Output Excel file to create.")
    parser.add_argument("--validate", type=Path, help="Optional reference workbook to compare BF:DO values against.")
    args = parser.parse_args()

    rows = build_output(args.input, args.output)
    print(f"Created {args.output} with {rows} data row(s).")

    if args.validate:
        mismatches = validate_output(args.output, args.validate)
        if mismatches:
            print("Validation failed:")
            for mismatch in mismatches:
                print(f"  - {mismatch}")
            return 1
        print("Validation passed: output BF:DO matches the reference workbook.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
